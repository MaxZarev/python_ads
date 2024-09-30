from __future__ import annotations

from openpyxl import Workbook, load_workbook
import os

from openpyxl.worksheet.worksheet import Worksheet
from config.config import excel_path

class Excel:
    def __init__(self):
        if not os.path.exists(excel_path):
            table = Workbook()
            table.active["A1"] = "Address"
            table.active["B1"] = "Password"
            table.active["C1"] = "Profile Number"
            table.active["D1"] = "Seed"
            table.save(excel_path)
        else:
            table = load_workbook(excel_path)
        self.file = excel_path
        self.table = table
        self.sheet: Worksheet = self.table.active

    def add_row(self, values: list) -> None:
        self.sheet.append(values)
        self.table.save(self.file)

    def add_cell(self, row_name: str, column_name: str, value: str) -> None:
        col_num = self.find_col(column_name)
        if not col_num:
            raise ValueError(f"Column with name '{column_name}' not found.")

        row_num = self.find_row(row_name)
        if not row_num:
            self.add_row([row_name])
            row_num = self.find_row(row_name)

        self.sheet.cell(row=row_num, column=col_num, value=value)
        self.table.save(self.file)

    def find_col(self, column_name: str) -> int:
        for row in self.sheet.iter_rows(max_row=1):
            for cell in row:
                if cell.value == column_name:
                    return cell.column
        return 0

    def find_row(self, row_name: str) -> int:
        for row in self.sheet.iter_rows(min_row=2, max_col=3):
            for cell in row:
                if cell.value == row_name:
                    return cell.row
        return 0

    def get_cell(self, row_name: str, column_name: str) -> str | None:
        col_num = self.find_col(column_name)
        row_num = self.find_row(row_name)

        if not row_num or not col_num:
            return None

        return self.sheet.cell(row=row_num, column=col_num).value

    def get_column(self, column_name: str) -> list[str]:
        col_num = self.find_col(column_name)
        if not col_num:
            raise ValueError(f"Column with name '{column_name}' not found.")
        column_values = []
        for raw in self.sheet.iter_cols(min_col=col_num, max_col=col_num, min_row=2):
            for cell in raw:
                column_values.append(cell.value)

        return column_values

    def get_row(self, row_name: str) -> list[str]:
        row_num = self.find_row(row_name)
        if not row_num:
            raise ValueError(f"Row with name '{row_name}' not found.")
        row_values = []
        for raw in self.sheet.iter_rows(min_row=row_num, max_row=row_num):
            for cell in raw:
                row_values.append(cell.value)

        return row_values




if __name__ == '__main__':
    excel = Excel("walelts.xlsx")
    # excel.add_cell("0x123", "Profile Number", "777")
    password = excel.get_cell("0x124", "Password")
    print(password)


# if __name__ == '__main__':
#     excel = Excel("okx_withdrawals.xlsx")
